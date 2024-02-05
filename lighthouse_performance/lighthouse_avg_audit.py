import datetime
import json
import os
import time
import subprocess
import csv

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from datetime import datetime
from features.page_objects.base_page_object import BasePage
from lighthouse_performance.lighthouse_report import LightHelper


class LighthouseAvgAudit(BasePage):
    def __init__(self, context):
        BasePage.__init__(self, context.browser)

    def run_audit(self, fcp_threshold, lcp_threshold, si_threshold, performance_threshold):
        helper = LightHelper(self, )

        url = self.browser.current_url

        url_ = f"{helper.set_filename(url)}"
        print(url_)

        # Create a list to store the file names
        json_files = []

        name = "light_report"
        getdate = datetime.now().strftime("%d-%m-%y")
        time_ = datetime.now().strftime("%H:%M:%S")

        date_time = getdate + '_' + time_

        # Set the Lighthouse options
        mobile_preset = 'desktop'
        chrome_flags = "--no-sandbox --headless --disable-gpu --disable-storage-reset --user-data-dir"
        port = '--port=9222'
        max_wait = '--max-wait-for-load=60000'

        relative_path = os.path.abspath(os.path.dirname(__file__))

        json_filename__ = os.path.join(relative_path, name)

        # Clear out JSON files
        for json_file in json_files:
            if os.path.exists(json_file):
                os.remove(json_file)
                print(f"{json_file} removed before run if exist")
            else:
                print(f"{json_file} JSON file now removed before audit")

        # prints parent window title
        print("Parent window title: " + self.browser.title)

        # get current window handle
        parent_window = self.browser.current_window_handle

        # get first child window
        child_window = self.browser.window_handles

        for w in child_window:
            # switch focus to child window
            if w != parent_window:
                self.browser.switch_to.window(w)
            break
        time.sleep(0.9)
        print("Child window title: " + self.browser.title)

        # Initialize lists to store metric values
        fcp_values = []
        lcp_values = []
        si_values = []
        performance_values = []

        # Set threshold
        # threshold set on step level

        # Run subprocess and collect metrics three times

        for i in range(1):
            # Create unique JSON filename for each iteration
            json_filename = json_filename__ + f"{i}output.json"

            subprocess.run(
                f'lighthouse {url} {port} {chrome_flags} --preset={mobile_preset} '
                f'{max_wait} --output=json --output-path=' + json_filename,
                shell=True
            )

            print("lighthouse run completed")

            time.sleep(60)

            # Open the JSON file and start processing it
            with open(json_filename, encoding="utf8") as json_data:
                loaded_json = json.load(json_data)

            # Extract performance metrics
            fcp_ms = float(loaded_json["audits"]["first-contentful-paint"]["numericValue"])
            lcp_ms = float(loaded_json["audits"]["largest-contentful-paint"]["numericValue"])
            si_ms = float(loaded_json["audits"]["speed-index"]["numericValue"])

            # Extract performance metrics
            try:
                performance_percentage = float(loaded_json["categories"]["performance"]["score"])
            except TypeError:
                print(f'No performance score for URL {url}, setting to 0.')
                performance_percentage = 0.0

            # Convert metrics from milliseconds to seconds
            fcp_sec = fcp_ms / 1000
            lcp_sec = lcp_ms / 1000
            si_sec = si_ms / 1000
            performance = performance_percentage * 100

            # Append metric values to the respective lists
            fcp_values.append(fcp_sec)
            lcp_values.append(lcp_sec)
            si_values.append(si_sec)
            performance_values.append(performance)

            # Store the file name
            json_files.append(json_filename)

            # Calculate average metrics
            average_fcp = sum(fcp_values) / len(fcp_values)
            average_lcp = sum(lcp_values) / len(lcp_values)
            average_si = sum(si_values) / len(si_values)
            average_performance = sum(performance_values) / len(performance_values)

            # Check if any thresholds are exceeded
            fcp_exceed_sec = max(0, average_fcp - fcp_threshold)
            lcp_exceed_sec = max(0, average_lcp - lcp_threshold)
            si_exceed_sec = max(0, average_si - si_threshold)
            perf_exceed_sec = max(0, average_performance - performance_threshold)

            # Save metrics to results.csv
            csv_output_file = os.path.join(relative_path, 'result.csv')

            # Check if the CSV file exists
            file_exists = os.path.isfile(csv_output_file)

            # Write the CSV file with headers if it does not exist or append new rows if it exists
            with open(csv_output_file, 'a', newline='') as csv_file:
                fieldnames = ['url_id', 'url', 'FCP_Threshold', 'LCP_Threshold', 'SI_Threshold', 'Perf_Threshold',
                              'FCP_Result', 'LCP_Result', 'SI_Result', 'Perf_Result', 'FCP_Exceeded', 'LCP_Exceeded',
                              'SI_Exceed', 'Perf_Exceed', 'date_time']
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

                # Write headers if the file does not exist
                if not file_exists:
                    writer.writeheader()

                # Write a new row with the average metrics in seconds
                writer.writerow({
                    'url_id': url,
                    'url': url_,
                    'FCP_Threshold': fcp_threshold,
                    'LCP_Threshold': lcp_threshold,
                    'SI_Threshold': si_threshold,
                    'Perf_Threshold': performance_threshold,
                    'FCP_Result': f"{average_fcp:.2f} s",
                    'SI_Result': f"{average_si:.2f} s",
                    'LCP_Result': f"{average_lcp:.2f} s",
                    'Perf_Result': f"{average_performance:.2f}",
                    'FCP_Exceeded': f"{fcp_exceed_sec:.2f}",
                    'LCP_Exceeded': f"{lcp_exceed_sec:.2f}",
                    'SI_Exceed': f"{si_exceed_sec:.2f}",
                    'Perf_Exceed': perf_exceed_sec,
                    'date_time': date_time
                })

            print("Lighthouse metrics saved to:", csv_output_file)

            # Remove the JSON file
            os.remove(json_filename)
            print("JSON file removed:", json_filename)

            # Remove the JSON files
            for json_file in json_files:
                if os.path.exists(json_file):
                    os.remove(json_file)
                    print(f"{json_file} removed")
                else:
                    print(f"{json_file} JSON file now removed after metrics extracted")

            print("Audit Done")

            # Convert the CSV to Excel
            df = pd.read_csv(csv_output_file)
            xlsx_output_file = os.path.join(relative_path, 'results.xlsx')
            df.to_excel(xlsx_output_file, index=False)

            # Define colors
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # Open the Excel workbook and select the active sheet
            wb = load_workbook(xlsx_output_file)
            ws = wb.active

            # Loop through rows and cells
            for row in ws.iter_rows(min_row=2):  # Assuming 1st row contains headers
                for cell in row:
                    column_title = ws.cell(row=1, column=cell.column).value
                    # Check conditions and fill color accordingly
                    if column_title in ['FCP_Threshold', 'LCP_Threshold', 'SI_Threshold', 'Perf_Threshold']:
                        cell.fill = yellow_fill
                    if column_title in ['FCP_Exceeded', 'LCP_Exceeded']:
                        if float(cell.value) > 0:
                            cell.fill = red_fill
                    if column_title in ['SI_Exceed', 'Perf_Exceed']:
                        if float(cell.value) < 0:
                            cell.fill = red_fill
                    elif column_title in ['FCP_Result', 'LCP_Result', 'SI_Result', 'Perf_Result']:
                        cell.fill = blue_fill

            # Save the workbook
            xlsx_output_file_ = os.path.join(relative_path, 'filtered_results.xlsx')
            wb.save(filename=xlsx_output_file_)

            # Remove the JSON file
            os.remove(xlsx_output_file)

        time.sleep(15)
        self.browser.switch_to.window(parent_window)
        self.browser.get(url)
