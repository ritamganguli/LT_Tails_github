import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import os


def parse_feature_file(file_path, f=3):
    """Parse the given feature file and extract scenarios, their steps, and tags within 'f' lines above the scenario,
    and specifically capture @nightly and @broken tags in separate columns."""
    with open(file_path, 'r') as file:
        lines = file.readlines()

    scenarios = []
    current_scenario = None
    steps = []
    buffer = []  # Buffer to hold last 'f' lines

    for line in lines:
        stripped_line = line.strip()

        # Check if the current line is a new Scenario
        if stripped_line.startswith('Scenario:'):
            if current_scenario:
                # Process the previous scenario
                current_scenario['Steps'] = '\n'.join(steps)
                # Extract @ tags from buffer
                current_scenario['Tags'] = ', '.join(tag for tag in buffer if tag.startswith('@'))
                # Extract @nightly tags
                current_scenario['Nightly'] = ', '.join(tag for tag in buffer if tag.startswith('@nightly'))
                # Extract @broken tags
                current_scenario['Broken'] = ', '.join(tag for tag in buffer if tag.startswith('@broken'))
                scenarios.append(current_scenario)

            # Start a new scenario
            current_scenario = {'Scenario': stripped_line.split(':', 1)[1].strip()}
            steps = []
            buffer = []  # Reset buffer for next scenario

        elif stripped_line.startswith(('Given ', 'When ', 'And ', 'Then ')):
            # If it's a step line, add it to the steps
            steps.append(stripped_line)
        else:
            # Update buffer, keep only last 'f' lines
            buffer.append(stripped_line)
            if len(buffer) > f:
                buffer.pop(0)

    # Process the last scenario outside the loop
    if current_scenario:
        current_scenario['Steps'] = '\n'.join(steps)
        # Extract @ tags from buffer for the last scenario
        current_scenario['Tags'] = ', '.join(tag for tag in buffer if tag.startswith('@'))
        # Extract @nightly tags for the last scenario
        current_scenario['Nightly'] = ', '.join(tag for tag in buffer if tag.startswith('@nightly'))
        # Extract @broken tags for the last scenario
        current_scenario['Broken'] = ', '.join(tag for tag in buffer if tag.startswith('@broken'))
        scenarios.append(current_scenario)

    return scenarios


def process_directory(directory_path):
    """Process each feature file in the given directory."""
    all_scenarios = []

    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.feature'):
                file_path = os.path.join(root, file)
                scenarios = parse_feature_file(file_path)
                all_scenarios.extend(scenarios)

    return all_scenarios


# Assuming this code is run in a standard Python environment where __file__ is defined
current_path = os.path.dirname(os.path.abspath(__file__))
print("first print" + current_path)
features_folder = "../features"  # The folder under the root path containing feature files
features_path = os.path.join(current_path, features_folder)
print("second print" + features_path)

# Process the directory and extract scenarios and steps
extracted_data = process_directory(features_path)

# Convert the extracted data to a DataFrame
extracted_df = pd.DataFrame(extracted_data)

# Save to CSV
output_csv_path = os.path.join(current_path, 'extracted_scenarios.csv')
extracted_df.to_csv(output_csv_path, index=False)


def find_duplicates(df):
    """Find duplicates across all columns in the dataframe."""
    duplicates = defaultdict(list)
    for col in df.columns:
        for word in df[col]:
            duplicates[word].append(col)
    return {word: cols for word, cols in duplicates.items() if len(cols) > 1}


# Load CSV file
file_path = 'extracted_scenarios.csv'
data = pd.read_csv(file_path)

# Find duplicates across all columns
duplicates = find_duplicates(data)

# Create an Excel file with color coding
wb = Workbook()
ws = wb.active

# Define colors for duplicates
colors = ["FFFF00", "FF0000", "00FF00", "0000FF", "FF00FF"]  # Yellow, Red, Green, Blue, Magenta
color_index = 0
fill_colors = {}

# Assign colors to duplicates
for word, _ in duplicates.items():
    fill_colors[word] = PatternFill(start_color=colors[color_index % len(colors)],
                                    end_color=colors[color_index % len(colors)],
                                    fill_type='solid')
    color_index += 1

# Write data to Excel with color coding for duplicates
for row_index, row in enumerate(data.itertuples(index=False), start=1):
    for col_index, word in enumerate(row, start=1):
        cell = ws.cell(row_index, col_index, word)
        if word in duplicates:
            cell.fill = fill_colors[word]

# Save the Excel file
output_file = 'analysis_output.xlsx'
wb.save(output_file)

print(f"Output saved to {output_file}")
